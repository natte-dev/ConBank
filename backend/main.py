"""
API FastAPI — Sistema de Conciliação de Fornecedores
SQLAlchemy 2.x  |  FastAPI 0.109  |  Pydantic v2
"""
import io
import os
import logging
import traceback
from contextlib import asynccontextmanager
from datetime import datetime
from decimal import Decimal
from typing import Optional

from fastapi import Depends, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from conciliacao_intel import conciliar_todos_fornecedores_inteligente
from consolidador import consolidar_todos_fornecedores
from database import get_db, init_db
from models import ArquivoImportado, ConciliacaoInterna, Divergencia, Fornecedor, LancamentoFornecedor
from parser import calcular_hash_arquivo, parsear_arquivo_razao

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CORS — separar múltiplas origens por vírgula no env
# Ex: ALLOWED_ORIGINS=https://frontend.easypanel.host,https://outro.com
# Default "*" só para facilitar primeiro deploy; restringir em produção real.
# ---------------------------------------------------------------------------
_raw_origins = os.getenv("ALLOWED_ORIGINS", "*")
CORS_ORIGINS: list[str] = [o.strip() for o in _raw_origins.split(",") if o.strip()]


# ---------------------------------------------------------------------------
# Lifespan — startup / shutdown (substitui @app.on_event depreciado)
# ---------------------------------------------------------------------------
@asynccontextmanager
async def lifespan(_app: FastAPI):
    """Inicializa o banco na subida e libera recursos na parada."""
    init_db()
    yield
    # Nada a fechar explicitamente — SQLAlchemy drena o pool sozinho


# ---------------------------------------------------------------------------
# Aplicação
# ---------------------------------------------------------------------------
app = FastAPI(
    title="Sistema de Conciliação de Fornecedores",
    description="API para conciliação interna de contas a pagar",
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================================
# UTILITÁRIOS
# ============================================================================

def _converter_data_br(valor) -> Optional[datetime]:
    """Converte string DD/MM/YYYY → datetime; passa datetime sem alterar."""
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor
    try:
        return datetime.strptime(str(valor), "%d/%m/%Y")
    except (ValueError, TypeError):
        return None


# ============================================================================
# ROTAS BÁSICAS
# ============================================================================

@app.get("/")
async def root():
    return {"message": "Sistema de Conciliação de Fornecedores", "version": "1.0.0", "status": "online"}


@app.get("/health")
async def health_check(db: Session = Depends(get_db)):
    """
    Healthcheck — usado pelo EasyPanel e pelo HEALTHCHECK do Dockerfile.
    Retorna 200 se o banco está acessível, 503 caso contrário.
    """
    try:
        db.execute(text("SELECT 1"))
        return {"status": "healthy", "database": "connected"}
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail={"status": "unhealthy", "error": str(exc)},
        )


# ============================================================================
# UPLOAD E PROCESSAMENTO
# ============================================================================

@app.post("/upload")
async def upload_arquivo(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Upload e processamento do arquivo PDF/ZIP do Razão de Fornecedores."""
    try:
        conteudo = await file.read()
        hash_arquivo = calcular_hash_arquivo(conteudo)

        # Idempotência: rejeita duplicatas
        if db.query(ArquivoImportado).filter(ArquivoImportado.hash_arquivo == hash_arquivo).first():
            raise HTTPException(status_code=400, detail="Arquivo já foi importado anteriormente")

        arquivo = ArquivoImportado(
            nome_arquivo=file.filename,
            hash_arquivo=hash_arquivo,
            status="PROCESSANDO",
        )
        db.add(arquivo)
        db.commit()
        db.refresh(arquivo)

        try:
            dados = parsear_arquivo_razao(conteudo)
            dados = consolidar_todos_fornecedores(dados)

            arquivo.data_inicio   = _converter_data_br(dados.get("periodo_inicio"))
            arquivo.data_fim      = _converter_data_br(dados.get("periodo_fim"))
            arquivo.empresa       = dados.get("empresa")
            arquivo.cnpj_empresa  = dados.get("cnpj")
            arquivo.total_fornecedores = dados.get("total_fornecedores", len(dados["fornecedores"]))
            arquivo.total_lancamentos  = dados.get(
                "total_lancamentos",
                sum(len(f.get("lancamentos", [])) for f in dados["fornecedores"]),
            )

            logger.info("💾 Inserindo %d fornecedores…", len(dados["fornecedores"]))

            for idx, forn_data in enumerate(dados["fornecedores"], 1):
                if idx % 50 == 0:
                    logger.info("   Processados %d / %d", idx, len(dados["fornecedores"]))

                saldo_anterior = Decimal(str(forn_data.get("saldo_anterior", 0)))
                total_credito  = Decimal(str(forn_data.get("total_credito", 0)))
                total_debito   = Decimal(str(forn_data.get("total_debito", 0)))

                fornecedor = Fornecedor(
                    arquivo_origem_id   = arquivo.id,
                    codigo_conta        = forn_data["codigo_conta"],
                    conta_contabil      = forn_data["conta_contabil"],
                    nome_fornecedor     = forn_data["nome_fornecedor"],
                    saldo_anterior      = saldo_anterior,
                    saldo_anterior_tipo = forn_data.get("saldo_anterior_tipo", ""),
                    total_debito        = total_debito,
                    total_credito       = total_credito,
                    saldo_final         = saldo_anterior + total_credito - total_debito,
                )
                db.add(fornecedor)
                db.flush()

                for lanc_data in forn_data.get("lancamentos", []):
                    vd     = Decimal(str(lanc_data["valor_debito"]))
                    vc     = Decimal(str(lanc_data["valor_credito"]))
                    saldo  = Decimal(str(lanc_data["saldo_apos_lancamento"]))

                    db.add(LancamentoFornecedor(
                        fornecedor_id         = fornecedor.id,
                        data_lancamento       = lanc_data["data_lancamento"],
                        lote                  = lanc_data.get("lote"),
                        historico             = lanc_data["historico"],
                        conta_partida         = lanc_data.get("conta_partida"),
                        valor_debito          = vd,
                        valor_credito         = vc,
                        saldo_apos_lancamento = saldo,
                        saldo_tipo            = lanc_data.get("saldo_tipo", ""),
                        tipo_operacao         = lanc_data["tipo_operacao"],
                        numero_nf             = lanc_data.get("numero_nf"),
                        cnpj_historico        = lanc_data.get("cnpj_historico"),
                        valor_saldo           = vc if lanc_data["tipo_operacao"] == "COMPRA" else Decimal("0"),
                    ))

            db.commit()
            logger.info("✅ Dados salvos no banco.")

            logger.info("🔄 Iniciando conciliação inteligente…")
            conciliar_todos_fornecedores_inteligente(db, arquivo.id)

            for forn in db.query(Fornecedor).filter(Fornecedor.arquivo_origem_id == arquivo.id).all():
                forn.valor_a_pagar = forn.total_credito - forn.total_debito
                if abs(forn.valor_a_pagar) <= Decimal("0.01"):
                    forn.status_pagamento = "QUITADO"
                elif forn.valor_a_pagar < 0:
                    forn.status_pagamento = "ADIANTADO"
                else:
                    forn.status_pagamento = "EM_ABERTO"

            arquivo.status = "CONCLUIDO"
            db.commit()
            logger.info("✅ Processamento concluído para arquivo_id=%d", arquivo.id)

            return {
                "success": True,
                "arquivo_id": arquivo.id,
                "message": "Arquivo processado com sucesso",
                "dados": {
                    "total_fornecedores": arquivo.total_fornecedores,
                    "total_lancamentos":  arquivo.total_lancamentos,
                    "periodo_inicio": arquivo.data_inicio.isoformat() if arquivo.data_inicio else None,
                    "periodo_fim":    arquivo.data_fim.isoformat()    if arquivo.data_fim    else None,
                },
            }

        except Exception as exc:
            logger.error("❌ Erro durante processamento:\n%s", traceback.format_exc())
            arquivo.status = "ERRO"
            arquivo.mensagem_erro = str(exc)
            db.commit()
            raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {exc}")

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# ============================================================================
# CONSULTAS
# ============================================================================

@app.get("/arquivos")
async def listar_arquivos(db: Session = Depends(get_db)):
    arquivos = db.query(ArquivoImportado).order_by(ArquivoImportado.created_at.desc()).all()
    return [
        {
            "id":                  arq.id,
            "nome_arquivo":        arq.nome_arquivo,
            "status":              arq.status,
            "total_fornecedores":  arq.total_fornecedores,
            "total_lancamentos":   arq.total_lancamentos,
            "periodo_inicio":      arq.data_inicio.isoformat() if arq.data_inicio else None,
            "periodo_fim":         arq.data_fim.isoformat()    if arq.data_fim    else None,
            "created_at":          arq.created_at.isoformat(),
        }
        for arq in arquivos
    ]


@app.get("/resumo/{arquivo_id}")
async def obter_resumo(arquivo_id: int, db: Session = Depends(get_db)):
    arquivo = db.query(ArquivoImportado).filter(ArquivoImportado.id == arquivo_id).first()
    if not arquivo:
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")

    fornecedores = db.query(Fornecedor).filter(Fornecedor.arquivo_origem_id == arquivo_id).all()

    return {
        "arquivo": {
            "id":            arquivo.id,
            "nome":          arquivo.nome_arquivo,
            "periodo_inicio": arquivo.data_inicio.isoformat() if arquivo.data_inicio else None,
            "periodo_fim":    arquivo.data_fim.isoformat()    if arquivo.data_fim    else None,
        },
        "estatisticas": {
            "total_fornecedores":        len(fornecedores),
            "total_lancamentos":         arquivo.total_lancamentos,
            "fornecedores_quitados":     sum(1 for f in fornecedores if f.status_pagamento == "QUITADO"),
            "fornecedores_em_aberto":    sum(1 for f in fornecedores if f.status_pagamento == "EM_ABERTO"),
            "fornecedores_adiantados":   sum(1 for f in fornecedores if f.status_pagamento == "ADIANTADO"),
            "fornecedores_com_divergencia": sum(1 for f in fornecedores if f.divergencia_calculo),
            "valor_total_a_pagar":       float(sum(f.valor_a_pagar or 0 for f in fornecedores)),
        },
    }


@app.get("/fornecedores")
async def listar_fornecedores(
    arquivo_id: int,
    status: Optional[str] = None,
    tem_parciais: Optional[bool] = Query(None, description="Filtrar fornecedores com NFs parcialmente pagas"),
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
):
    q = db.query(Fornecedor).filter(Fornecedor.arquivo_origem_id == arquivo_id)

    if status:
        q = q.filter(Fornecedor.status_pagamento == status)
    if tem_parciais is True:
        q = q.filter(Fornecedor.qtd_nfs_parciais > 0)
    elif tem_parciais is False:
        q = q.filter(Fornecedor.qtd_nfs_parciais == 0)

    total = q.count()
    fornecedores = q.order_by(Fornecedor.valor_a_pagar.desc()).offset(skip).limit(limit).all()

    return {
        "total": total,
        "fornecedores": [
            {
                "id":               f.id,
                "codigo_conta":     f.codigo_conta,
                "conta_contabil":   f.conta_contabil,
                "nome_fornecedor":  f.nome_fornecedor,
                "total_credito":    float(f.total_credito or 0),
                "total_debito":     float(f.total_debito or 0),
                "saldo_final":      float(f.saldo_final or 0),
                "valor_a_pagar":    float(f.valor_a_pagar or 0),
                "status_pagamento": f.status_pagamento,
                "qtd_nfs_pendentes": f.qtd_nfs_pendentes,
                "qtd_nfs_parciais":  f.qtd_nfs_parciais,
                "divergencia_calculo": f.divergencia_calculo,
            }
            for f in fornecedores
        ],
    }


@app.get("/fornecedores/{fornecedor_id}")
async def obter_fornecedor_detalhado(fornecedor_id: int, db: Session = Depends(get_db)):
    fornecedor = db.query(Fornecedor).filter(Fornecedor.id == fornecedor_id).first()
    if not fornecedor:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado")

    lancamentos = (
        db.query(LancamentoFornecedor)
        .filter(LancamentoFornecedor.fornecedor_id == fornecedor_id)
        .order_by(LancamentoFornecedor.data_lancamento)
        .all()
    )

    compras_pendentes = [l for l in lancamentos if l.tipo_operacao == "COMPRA" and (l.valor_saldo or 0) > 0]

    return {
        "fornecedor": {
            "id":                 fornecedor.id,
            "codigo_conta":       fornecedor.codigo_conta,
            "conta_contabil":     fornecedor.conta_contabil,
            "nome_fornecedor":    fornecedor.nome_fornecedor,
            "cnpj":               fornecedor.cnpj,
            "saldo_anterior":     float(fornecedor.saldo_anterior or 0),
            "total_credito":      float(fornecedor.total_credito or 0),
            "total_debito":       float(fornecedor.total_debito or 0),
            "saldo_final":        float(fornecedor.saldo_final or 0),
            "valor_a_pagar":      float(fornecedor.valor_a_pagar or 0),
            "status_pagamento":   fornecedor.status_pagamento,
            "divergencia_calculo": fornecedor.divergencia_calculo,
        },
        "compras_pendentes": [
            {
                "id":               c.id,
                "data_lancamento":  c.data_lancamento.isoformat() if c.data_lancamento else None,
                "numero_nf":        c.numero_nf,
                "historico":        c.historico,
                "valor_total":      float(c.valor_credito),
                "valor_pago_parcial": float(c.valor_pago_parcial or 0),
                "valor_saldo":      float(c.valor_saldo or 0),
                "status_pagamento": c.status_pagamento,
            }
            for c in compras_pendentes
        ],
        "todos_lancamentos": [
            {
                "id":            l.id,
                "data":          l.data_lancamento.isoformat() if l.data_lancamento else None,
                "lote":          l.lote,
                "historico":     l.historico,
                "tipo_operacao": l.tipo_operacao,
                "valor_debito":  float(l.valor_debito or 0),
                "valor_credito": float(l.valor_credito or 0),
                "saldo_apos":    float(l.saldo_apos_lancamento or 0),
            }
            for l in lancamentos
        ],
    }


@app.get("/divergencias")
async def listar_divergencias(arquivo_id: int, db: Session = Depends(get_db)):
    divergencias = (
        db.query(Divergencia)
        .join(Fornecedor)
        .filter(
            Fornecedor.arquivo_origem_id == arquivo_id,
            Divergencia.resolvido.is_(False),
        )
        .all()
    )
    return [
        {
            "id":            d.id,
            "fornecedor_id": d.fornecedor_id,
            "tipo":          d.tipo,
            "severidade":    d.severidade,
            "descricao":     d.descricao,
            "diferenca":     float(d.diferenca or 0),
            "created_at":    d.created_at.isoformat(),
        }
        for d in divergencias
    ]


# ============================================================================
# EXPORT
# ============================================================================

@app.get("/export/excel/{arquivo_id}")
async def exportar_excel(
    arquivo_id: int,
    tipo: str = Query("completo", pattern="^(completo|em_aberto|divergencias)$"),
    db: Session = Depends(get_db),
):
    """Exporta dados para Excel (.xlsx) em memória, sem gravar em disco."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    arquivo = db.query(ArquivoImportado).filter(ArquivoImportado.id == arquivo_id).first()
    if not arquivo:
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliação Fornecedores"

    headers = ["Código", "Conta Contábil", "Fornecedor", "CNPJ",
               "Total Compras", "Total Pagamentos", "Saldo a Pagar",
               "Status", "NFs Pendentes", "Divergência"]

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    q = db.query(Fornecedor).filter(Fornecedor.arquivo_origem_id == arquivo_id)
    if tipo == "em_aberto":
        q = q.filter(Fornecedor.status_pagamento == "EM_ABERTO")
    elif tipo == "divergencias":
        q = q.filter(Fornecedor.divergencia_calculo.is_(True))

    for row, f in enumerate(q.order_by(Fornecedor.valor_a_pagar.desc()).all(), 2):
        ws.cell(row=row, column=1,  value=f.codigo_conta)
        ws.cell(row=row, column=2,  value=f.conta_contabil)
        ws.cell(row=row, column=3,  value=f.nome_fornecedor)
        ws.cell(row=row, column=4,  value=f.cnpj)
        ws.cell(row=row, column=5,  value=float(f.total_credito or 0))
        ws.cell(row=row, column=6,  value=float(f.total_debito or 0))
        ws.cell(row=row, column=7,  value=float(f.valor_a_pagar or 0))
        ws.cell(row=row, column=8,  value=f.status_pagamento)
        ws.cell(row=row, column=9,  value=f.qtd_nfs_pendentes)
        ws.cell(row=row, column=10, value="Sim" if f.divergencia_calculo else "Não")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=conciliacao_{tipo}.xlsx"},
    )


# ============================================================================
# ENTRYPOINT LOCAL (não usado pelo Docker — o CMD do Dockerfile chama uvicorn)
# ============================================================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
